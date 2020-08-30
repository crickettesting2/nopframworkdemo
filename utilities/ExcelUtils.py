import openpyxl

def getRowcount(file,sheetName): #no of rows present in a sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):          #no of column present in a sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(file,sheetName,rownumber,columnnumber):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownumber,column=columnnumber).value

def writeData(file,sheetName,rownumber,columnnumber,data):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownumber,column=columnnumber).value=data
    workbook.save(file)
